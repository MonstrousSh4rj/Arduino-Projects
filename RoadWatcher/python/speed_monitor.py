import serial, serial.tools.list_ports
import tkinter as tk
from tkinter import font as tkfont
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# ── Settings ─────────────────────────────────
SPEED_LIMIT = 50
FINE_RS     = 5000
EXCEL_FILE  = "speed_log.xlsx"

SENDER_EMAIL    = "yourgmail@gmail.com"
SENDER_PASSWORD = "your_app_password"
RECEIVER_EMAIL  = "receiver@gmail.com"
# ─────────────────────────────────────────────

def send_fine_email(speed, fine, date, time):
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"Traffic Fine Notice — Overspeeding Rs.{fine}"
        msg["From"]    = SENDER_EMAIL
        msg["To"]      = RECEIVER_EMAIL
        body = f"""
Dear Car Owner,

Your vehicle has been detected OVERSPEEDING.

━━━━━━━━━━━━━━━━━━━━━━━━━
  FINE NOTICE
━━━━━━━━━━━━━━━━━━━━━━━━━
  Date        : {date}
  Time        : {time}
  Speed       : {speed} km/h
  Speed Limit : {SPEED_LIMIT} km/h
  Fine Amount : Rs. {fine}
━━━━━━━━━━━━━━━━━━━━━━━━━

Please pay within 7 days.

Regards,
Speed Monitoring System
        """
        msg.attach(MIMEText(body, "plain"))
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.sendmail(SENDER_EMAIL, RECEIVER_EMAIL, msg.as_string())
        server.quit()
        return True
    except Exception as e:
        print(f"Email error: {e}")
        return False

def setup_excel():
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        return wb, wb.active
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Speed Log"
    ws.append(["#", "Date", "Time", "Speed (km/h)", "Status", "Fine (Rs.)", "Email Sent"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2D3561")
        cell.alignment = Alignment(horizontal="center")
    for col, w in zip("ABCDEFG", [6,14,12,16,16,14,12]):
        ws.column_dimensions[col].width = w
    wb.save(EXCEL_FILE)
    return wb, ws

def log_speed(wb, ws, spd, status, email_sent=False):
    n = ws.max_row
    now = datetime.now()
    fine = FINE_RS if status == "OverSpeed" else 0
    ws.append([n, now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"),
               round(spd,1), status, fine, "Yes" if email_sent else "—"])
    if status == "OverSpeed":
        for cell in ws[ws.max_row]:
            cell.fill = PatternFill("solid", fgColor="FFD5D5")
            cell.font = Font(color="8B0000")
    wb.save(EXCEL_FILE)
    return n, fine, now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S")

def find_port():
    for p in serial.tools.list_ports.comports():
        if any(k in p.description for k in ["Arduino","CH340","USB Serial","ttyUSB"]):
            return p.device
    return "COM3"

PORT = find_port()
try:
    ser = serial.Serial(PORT, 9600, timeout=1)
except Exception as e:
    print(f"Could not open port {PORT}: {e}")
    exit()

wb, ws = setup_excel()

root = tk.Tk()
root.title("RoadWatcher — Speed Monitor")
root.geometry("540x540")
root.configure(bg="#1a1a2e")
root.resizable(False, False)

BIG = tkfont.Font(family="Arial", size=54, weight="bold")
MED = tkfont.Font(family="Arial", size=13)
SML = tkfont.Font(family="Arial", size=11)

tk.Label(root, text="🚗 RoadWatcher", font=("Arial",18,"bold"),
         bg="#1a1a2e", fg="white").pack(pady=10)

speed_var = tk.StringVar(value="-- km/h")
speed_lbl = tk.Label(root, textvariable=speed_var, font=BIG,
                     bg="#1a1a2e", fg="#00d4ff")
speed_lbl.pack()

status_var = tk.StringVar(value="Waiting for car...")
tk.Label(root, textvariable=status_var, font=MED,
         bg="#1a1a2e", fg="#aaaaaa").pack(pady=4)

alert_var = tk.StringVar(value="")
alert_lbl = tk.Label(root, textvariable=alert_var, font=("Arial",15,"bold"),
                     bg="#1a1a2e", fg="white")
alert_lbl.pack()

email_var = tk.StringVar(value="")
email_lbl = tk.Label(root, textvariable=email_var, font=("Arial",11),
                     bg="#1a1a2e", fg="#aaaaaa")
email_lbl.pack()

stats = tk.Frame(root, bg="#0f0f1e"); stats.pack(fill="x", padx=20, pady=8)
total_var = tk.StringVar(value="Total cars: 0")
fine_var  = tk.StringVar(value="Fines issued: 0")
max_var   = tk.StringVar(value="Max speed: 0 km/h")
for v in [total_var, fine_var, max_var]:
    tk.Label(stats, textvariable=v, font=SML,
             bg="#0f0f1e", fg="#aaaaaa").pack(side="left", expand=True)

tk.Label(root, text="Recent detections", font=SML,
         bg="#1a1a2e", fg="#555").pack()
log_box = tk.Listbox(root, height=7, bg="#0f0f1e", fg="#ccc",
                     font=SML, selectbackground="#333", bd=0)
log_box.pack(fill="x", padx=20, pady=4)

tk.Label(root, text=f"Port: {PORT}  |  Limit: {SPEED_LIMIT} km/h  |  Fine: Rs.{FINE_RS}  |  Log: {EXCEL_FILE}",
         font=("Arial",9), bg="#1a1a2e", fg="#444").pack(side="bottom", pady=5)

total_count = fine_count = 0
max_spd = 0.0

def read_serial():
    global total_count, fine_count, max_spd
    try:
        if ser.in_waiting:
            line = ser.readline().decode("utf-8", errors="ignore").strip()
            if not line: root.after(100, read_serial); return

            if line.startswith("SPEED:"):
                spd = float(line.split(":")[1])
                speed_var.set(f"{spd:.1f} km/h")
                status_var.set("Car detected!")
                total_count += 1
                if spd > max_spd: max_spd = spd
                total_var.set(f"Total cars: {total_count}")
                max_var.set(f"Max speed: {max_spd:.1f} km/h")

            elif line.startswith("ALERT:"):
                msg = line.split(":")[1]
                try: spd = float(speed_var.get().split()[0])
                except: spd = 0.0

                if msg == "OverSpeed":
                    email_var.set("Sending fine email...")
                    root.update()
                    now = datetime.now()
                    email_sent = send_fine_email(
                        round(spd,1), FINE_RS,
                        now.strftime("%Y-%m-%d"),
                        now.strftime("%H:%M:%S")
                    )
                    row, fine, date, time = log_speed(wb, ws, spd, msg, email_sent)
                    alert_var.set("OVER SPEEDING — Fine issued!")
                    alert_lbl.config(fg="#ff4444")
                    speed_lbl.config(fg="#ff4444")
                    fine_count += 1
                    fine_var.set(f"Fines issued: {fine_count}")
                    if email_sent:
                        email_var.set(f"Email sent to {RECEIVER_EMAIL}")
                        email_lbl.config(fg="#00ff99")
                    else:
                        email_var.set("Email failed — check settings")
                        email_lbl.config(fg="#ff4444")
                    log_box.insert(0, f"#{row}  {time}  {speed_var.get()}  FINE Rs.{fine}  {'EMAIL SENT' if email_sent else 'NO EMAIL'}")
                    log_box.itemconfig(0, fg="#ff6666")
                else:
                    row, fine, date, time = log_speed(wb, ws, spd, msg)
                    alert_var.set("Normal speed")
                    alert_lbl.config(fg="#00ff99")
                    speed_lbl.config(fg="#00d4ff")
                    email_var.set("")
                    log_box.insert(0, f"#{row}  {time}  {speed_var.get()}  OK")

            elif "Waiting" in line:
                status_var.set("Waiting for car...")
                speed_var.set("-- km/h")
                alert_var.set("")
                email_var.set("")
                speed_lbl.config(fg="#00d4ff")
            elif "Searching" in line:
                status_var.set("Car entering... searching")

    except Exception as e:
        status_var.set(f"Serial error: {e}")
    root.after(100, read_serial)

read_serial()
root.mainloop()
